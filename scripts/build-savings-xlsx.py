#!/usr/bin/env python3
"""Build a 4-sheet xlsx report from test/results/token-savings.v2.json.

Sheets: Summary, Per-Prompt, Raw-JSON, Methodology.
"""
import json
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = REPO_ROOT / "test" / "results" / "token-savings.v2.json"
XLSX_PATH = REPO_ROOT / "test" / "results" / "token-savings.xlsx"

VARIANTS = [
    "tc-token-optimizer",
    "tc-mirror-minimal",
    "tc-mirror-glyphs",
    "tc-mirror-classical",
]

HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
NA_FILL = PatternFill("solid", start_color="E7E6E6")

GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")
RED_FILL = PatternFill("solid", start_color="FFC7CE")


def apply_header(row_cells):
    for c in row_cells:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_summary(wb, data):
    ws = wb.create_sheet("Summary")
    headers = [
        "variant",
        "fixtures_measured",
        "mean_savings_pct",
        "median_savings_pct",
        "mean_output_tokens",
        "mean_overhead_tokens",
    ]
    ws.append(headers)
    apply_header(ws[1])

    for variant in VARIANTS:
        savings = []
        outputs = []
        overheads = []
        for f in data["fixtures"]:
            v = f["variants"][variant]
            if v.get("expectedOutputTokens") is not None:
                outputs.append(v["expectedOutputTokens"])
            if v.get("savingsPct") is not None:
                savings.append(v["savingsPct"])
            if v.get("skillOverheadTokens") is not None:
                overheads.append(v["skillOverheadTokens"])

        if not savings:
            row = [variant, len(outputs), "n/a", "n/a", "n/a",
                   round(statistics.mean(overheads), 1) if overheads else "n/a"]
        else:
            row = [
                variant,
                len(outputs),
                round(statistics.mean(savings), 2),
                round(statistics.median(savings), 2),
                round(statistics.mean(outputs), 2) if outputs else "n/a",
                round(statistics.mean(overheads), 1) if overheads else "n/a",
            ]
        ws.append(row)

    # Style body
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            if cell.value == "n/a":
                cell.fill = NA_FILL
                cell.alignment = Alignment(horizontal="center")

    # Widths
    ws.column_dimensions["A"].width = 22
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"


def savings_label(value):
    if value is None:
        return "n/a"
    return value


def build_per_prompt(wb, data):
    ws = wb.create_sheet("Per-Prompt")
    headers = [
        "fixture_id",
        "prompt_text",
        "prompt_tokens",
        "en_prompt_tokens",
        "baseline_tokens",
    ]
    for v in VARIANTS:
        headers.append(f"{v}_tokens")
        headers.append(f"{v}_savings_%")
    ws.append(headers)
    apply_header(ws[1])

    for f in data["fixtures"]:
        row = [
            f["id"],
            f["promptText"],
            f["promptTokens"],
            f.get("enTranslatedPromptTokens") if f.get("enTranslatedPromptTokens") is not None else "n/a",
            f["baselineOutputTokens"],
        ]
        for v in VARIANTS:
            vd = f["variants"][v]
            row.append(vd["expectedOutputTokens"] if vd["expectedOutputTokens"] is not None else "n/a")
            row.append(vd["savingsPct"] if vd["savingsPct"] is not None else "n/a")
        ws.append(row)

    # Style body
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            if cell.value == "n/a":
                cell.fill = NA_FILL
                cell.alignment = Alignment(horizontal="center")
        # prompt_text wrap
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Conditional formatting on savings_% columns (every other variant col, starting at col 7)
    # variants begin at column 6 (tokens) / 7 (savings)
    savings_cols = []
    for i, v in enumerate(VARIANTS):
        col_idx = 6 + i * 2 + 1  # savings col
        savings_cols.append(col_idx)

    last_row = ws.max_row
    for col_idx in savings_cols:
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThanOrEqual", formula=["80"], fill=GREEN_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="between", formula=["50", "79.999"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["50"], fill=RED_FILL))

    # Freeze first column AND header row
    ws.freeze_panes = "B2"


def build_raw_json(wb, data):
    ws = wb.create_sheet("Raw-JSON")
    pretty = json.dumps(data, ensure_ascii=False, indent=2)
    cell = ws.cell(row=1, column=1, value=pretty)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(name="Consolas", size=10)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 400


def build_methodology(wb, data):
    ws = wb.create_sheet("Methodology")
    ws.append(["Key", "Value"])
    apply_header(ws[1])

    rows = [
        ("Tokenizer", "cl100k_base (via js-tiktoken) — OpenAI's, NOT Anthropic's. Numbers are directional only."),
        ("Baseline output", "Constant 120 tokens (placeholder for a typical verbose unconstrained reply). NOT a live Claude measurement."),
        ("Expected output source", "Hand-written fixtures at test/fixtures/expected-outputs/<variant>/<prompt-id>.md. Currently only 01-bugfix exists for 3 of 4 variants. Other 7 prompts have null."),
        ("EN-translated baseline", "Tokens of the English twin prompt (08-pure-en-control is its own .en). Where no .en sibling exists, null."),
        ("Measured at", data.get("measuredAt", "n/a")),
        ("Limitations", "(a) Tokenizer mismatch (cl100k != Anthropic). (b) Hand-written expected outputs != live Claude replies — see FUTURE_WORK.md item 1. (c) Glyph rendering depends on terminal unicode support. (d) Skill picker selection is not exercised; this measures intended output."),
        ("Plan reference", "docs/superpowers/plans/2026-05-14-tc-mirror-variants.md Task 5"),
    ]
    for key, value in rows:
        ws.append([key, value])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(name="Arial", bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=r, column=2).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80


def main():
    if not JSON_PATH.exists():
        raise SystemExit(f"missing input: {JSON_PATH}")
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_summary(wb, data)
    build_per_prompt(wb, data)
    build_raw_json(wb, data)
    build_methodology(wb, data)

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"wrote {XLSX_PATH}")


if __name__ == "__main__":
    main()
